"""
app_atrasos.py - Dashboard de Atrasos Aeroportuarios
===================================================

Streamlit app para:
1. Fazer upload de arquivos RD_*.xlsx diarios
2. Extrair e consolidar automaticamente em base CSV
3. Visualizar e filtrar os dados

Instalacao:
    pip install -r requirements.txt

Execucao:
    streamlit run app_atrasos.py
"""

import datetime
import io
import os
from typing import Dict, List, Optional, Set, Tuple

import pandas as pd
import plotly.express as px
import streamlit as st


BASE_CSV = "base_atrasos.csv"
PAGE_TITLE = "Atrasos Aeroportuarios"
PAGE_ICON = "✈️"

COLUNAS = [
    "data",
    "icao",
    "aeroporto",
    "item",
    "tipo_ocorrencia",
    "movimento",
    "motivo_1",
    "minutos_motivo_1",
    "motivo_2",
    "minutos_motivo_2",
    "motivo_3",
    "minutos_motivo_3",
    "companhia",
    "numero_voo",
    "equipamento",
    "origem_destino",
    "af_aeroporto",
]

COLUNAS_MINUTOS = [
    "minutos_motivo_1",
    "minutos_motivo_2",
    "minutos_motivo_3",
]

PLOT_THEME = dict(
    paper_bgcolor="#0d1117",
    plot_bgcolor="#0d1117",
    font_color="#e6edf3",
    font_family="DM Sans",
    colorway=["#58a6ff", "#3fb950", "#d29922", "#f85149", "#bc8cff", "#39d353"],
)

CSS_CUSTOMIZADO = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=DM+Sans:wght@300;400;600;700&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.stApp { background-color: #0d1117; color: #e6edf3; }
section[data-testid="stSidebar"] {
    background-color: #161b22;
    border-right: 1px solid #30363d;
}

h1, h2, h3 { font-family: 'Space Mono', monospace !important; }

[data-testid="metric-container"] {
    background: #161b22;
    border: 1px solid #30363d;
    border-radius: 8px;
    padding: 16px 20px;
}
[data-testid="metric-container"] label {
    color: #8b949e !important;
    font-size: 0.75rem;
}
[data-testid="metric-container"] [data-testid="stMetricValue"] {
    color: #58a6ff !important;
    font-family: 'Space Mono', monospace !important;
    font-size: 1.8rem !important;
}
[data-testid="stMetricDelta"] { font-size: 0.75rem; }

.stButton > button {
    background: #238636;
    color: white;
    border: none;
    border-radius: 6px;
    font-family: 'DM Sans', sans-serif;
    font-weight: 600;
    padding: 0.5rem 1.2rem;
    transition: background 0.2s;
}
.stButton > button:hover { background: #2ea043; }

[data-testid="stFileUploader"] {
    background: #161b22;
    border: 1px dashed #388bfd;
    border-radius: 8px;
    padding: 1rem;
}

[data-testid="stDataFrame"] {
    border: 1px solid #30363d;
    border-radius: 8px;
}

.stTabs [data-baseweb="tab-list"] {
    background: #161b22;
    border-radius: 8px;
    gap: 4px;
}
.stTabs [data-baseweb="tab"] {
    color: #8b949e;
    font-weight: 600;
}
.stTabs [data-baseweb="tab"][aria-selected="true"] {
    color: #58a6ff !important;
    border-bottom: 2px solid #58a6ff;
    background: transparent;
}

[data-baseweb="select"] { background: #161b22; }
hr { border-color: #30363d; }

.badge {
    display: inline-block;
    background: #1f3a5f;
    color: #58a6ff;
    border: 1px solid #388bfd44;
    border-radius: 20px;
    padding: 2px 10px;
    font-size: 0.78rem;
    font-family: 'Space Mono', monospace;
    font-weight: 700;
    margin: 2px;
}
.badge-red {
    background: #3d1c1c;
    color: #f85149;
    border-color: #f8514944;
}
.badge-yellow {
    background: #3d2e0a;
    color: #d29922;
    border-color: #d2992244;
}
.badge-green {
    background: #1a3028;
    color: #3fb950;
    border-color: #3fb95044;
}
.section-header {
    font-family: 'Space Mono', monospace;
    font-size: 0.7rem;
    letter-spacing: 0.15em;
    color: #8b949e;
    text-transform: uppercase;
    margin-bottom: 0.5rem;
    border-bottom: 1px solid #30363d;
    padding-bottom: 4px;
}
</style>
"""


def configurar_pagina() -> None:
    st.set_page_config(
        page_title=PAGE_TITLE,
        page_icon=PAGE_ICON,
        layout="wide",
        initial_sidebar_state="expanded",
    )
    st.markdown(CSS_CUSTOMIZADO, unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# Extracao de dados do workbook
# -----------------------------------------------------------------------------

def normalizar_minutos(valor):
    if valor is None:
        return None
    if isinstance(valor, datetime.time):
        return valor.hour * 60 + valor.minute
    try:
        return int(valor)
    except (TypeError, ValueError):
        return None



def obter_nome_aeroporto(ws) -> str:
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and "Aeroporto" in cell:
                return cell.strip()
    return ""



def encontrar_linha_ocorrencias(ws) -> Optional[int]:
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=60, values_only=True), start=1):
        primeira_coluna = row[0] if row else None
        if primeira_coluna and isinstance(primeira_coluna, str) and "OCORRÊNCIAS" in primeira_coluna:
            return i
    return None



def extrair_de_bytes(file_bytes: bytes, data_str: str) -> pd.DataFrame:
    """Extrai dados de um arquivo xlsx em memoria. data_str no formato DD/MM/YYYY."""
    from openpyxl import load_workbook

    registros: List[dict] = []
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    try:
        abas_aeroporto = [nome_aba for nome_aba in wb.sheetnames if nome_aba.startswith("SB")]

        for icao in abas_aeroporto:
            ws = wb[icao]
            nome_aeroporto = obter_nome_aeroporto(ws)
            linha_header = encontrar_linha_ocorrencias(ws)

            if not linha_header:
                continue

            data_inicio = linha_header + 2
            for row in ws.iter_rows(
                min_row=data_inicio,
                max_row=data_inicio + 120,
                values_only=True,
            ):
                tipo = row[1] if len(row) > 1 else None
                if not tipo or not isinstance(tipo, str):
                    continue

                registros.append(
                    {
                        "data": data_str,
                        "icao": icao,
                        "aeroporto": nome_aeroporto,
                        "item": row[0] if len(row) > 0 else None,
                        "tipo_ocorrencia": tipo.strip(),
                        "movimento": row[2] if len(row) > 2 else None,
                        "motivo_1": row[3] if len(row) > 3 else None,
                        "minutos_motivo_1": normalizar_minutos(row[4] if len(row) > 4 else None),
                        "motivo_2": row[5] if len(row) > 5 else None,
                        "minutos_motivo_2": normalizar_minutos(row[6] if len(row) > 6 else None),
                        "motivo_3": row[7] if len(row) > 7 else None,
                        "minutos_motivo_3": normalizar_minutos(row[8] if len(row) > 8 else None),
                        "companhia": row[9] if len(row) > 9 else None,
                        "numero_voo": str(row[10]) if len(row) > 10 and row[10] else None,
                        "equipamento": row[11] if len(row) > 11 else None,
                        "origem_destino": row[12] if len(row) > 12 else None,
                        "af_aeroporto": row[13] if len(row) > 13 else None,
                    }
                )
    finally:
        wb.close()

    return pd.DataFrame(registros) if registros else pd.DataFrame(columns=COLUNAS)


# -----------------------------------------------------------------------------
# Persistencia e sessao
# -----------------------------------------------------------------------------

def data_valida(valor: str) -> bool:
    """Retorna True se o valor for uma data no formato DD/MM/YYYY."""
    if not isinstance(valor, str):
        return False
    try:
        datetime.datetime.strptime(valor.strip(), "%d/%m/%Y")
        return True
    except ValueError:
        return False


@st.cache_data

def carregar_base_csv(caminho: str) -> pd.DataFrame:
    if os.path.exists(caminho):
        df = pd.read_csv(caminho, dtype=str, encoding="utf-8-sig")

        if "data" in df.columns:
            df = df[df["data"].apply(data_valida)].reset_index(drop=True)

        for col in COLUNAS_MINUTOS:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        return df

    return pd.DataFrame(columns=COLUNAS)



def salvar_base_csv(df: pd.DataFrame, caminho: str) -> None:
    if "data" in df.columns:
        df = df[df["data"].apply(data_valida)].reset_index(drop=True)
    df.to_csv(caminho, index=False, encoding="utf-8-sig")



def inicializar_session_state() -> None:
    if "df_base" not in st.session_state:
        st.session_state.df_base = carregar_base_csv(BASE_CSV)

    if "datas_processadas" not in st.session_state:
        df_base = st.session_state.df_base
        st.session_state.datas_processadas = set(df_base["data"].unique()) if not df_base.empty else set()


# -----------------------------------------------------------------------------
# Sidebar
# -----------------------------------------------------------------------------

def processar_uploads(uploaded_files, datas_por_arquivo: Dict[str, datetime.date]) -> List[str]:
    novos_dfs = []
    log: List[str] = []

    for arquivo in uploaded_files:
        data_sel = datas_por_arquivo[arquivo.name]
        data_arq = data_sel.strftime("%d/%m/%Y")

        if data_arq in st.session_state.datas_processadas:
            log.append(f"**{arquivo.name}** - ja na base ({data_arq})")
            continue

        with st.spinner(f"Extraindo {arquivo.name}..."):
            df_novo = extrair_de_bytes(arquivo.read(), data_arq)

        if df_novo.empty:
            log.append(f"**{arquivo.name}** - nenhuma ocorrencia encontrada")
            continue

        novos_dfs.append(df_novo)
        st.session_state.datas_processadas.add(data_arq)
        log.append(f"**{arquivo.name}** - {len(df_novo)} ocorrencias ({data_arq})")

    if novos_dfs:
        df_combined = pd.concat([st.session_state.df_base] + novos_dfs, ignore_index=True)
        st.session_state.df_base = df_combined
        salvar_base_csv(df_combined, BASE_CSV)
        st.cache_data.clear()

    return log



def render_sidebar() -> None:
    with st.sidebar:
        st.markdown("## Atrasos Aeroportuarios")
        st.markdown('<div class="section-header">Importar Arquivos</div>', unsafe_allow_html=True)

        uploaded_files = st.file_uploader(
            "Arquivos RD_*.xlsx",
            type=["xlsx"],
            accept_multiple_files=True,
            help="Selecione a data de cada arquivo antes de processar.",
        )

        datas_por_arquivo: Dict[str, datetime.date] = {}
        if uploaded_files:
            st.markdown('<div class="section-header">Data de cada arquivo</div>', unsafe_allow_html=True)
            for arquivo in uploaded_files:
                datas_por_arquivo[arquivo.name] = st.date_input(
                    arquivo.name,
                    value=datetime.date.today(),
                    format="DD/MM/YYYY",
                    key=f"date_{arquivo.name}",
                )

            if st.button("Processar e Adicionar a Base", use_container_width=True):
                mensagens_log = processar_uploads(uploaded_files, datas_por_arquivo)
                for msg in mensagens_log:
                    st.markdown(msg)

        st.divider()
        render_sidebar_filtros()
        st.divider()
        render_sidebar_exportacao()



def render_sidebar_filtros() -> None:
    st.markdown('<div class="section-header">Filtros</div>', unsafe_allow_html=True)
    df_all = st.session_state.df_base.copy()

    if df_all.empty:
        st.info("Nenhum dado na base ainda.")
        return

    df_all["_ano"] = df_all["data"].astype(str).str[-4:]
    anos_disp = sorted(df_all["_ano"].dropna().unique().tolist(), reverse=True)
    anos_sel = st.multiselect("Ano", anos_disp, default=anos_disp, key="f_ano")

    df_all_filtrado = df_all[df_all["_ano"].isin(anos_sel)] if anos_sel else df_all
    datas_disp = sorted(df_all_filtrado["data"].dropna().unique().tolist())
    st.multiselect("Data", datas_disp, default=datas_disp, key="f_data")

    icaos_disp = sorted(df_all["icao"].dropna().unique().tolist())
    st.multiselect("Aeroporto (ICAO)", icaos_disp, default=icaos_disp, key="f_icao")

    tipos_disp = sorted(df_all["tipo_ocorrencia"].dropna().unique().tolist())
    st.multiselect("Tipo de Ocorrencia", tipos_disp, default=tipos_disp, key="f_tipo")

    movs_disp = sorted(df_all["movimento"].dropna().unique().tolist())
    st.multiselect("Movimento", movs_disp, default=movs_disp, key="f_mov")

    cias_disp = sorted(df_all["companhia"].dropna().unique().tolist())
    st.multiselect("Companhia", cias_disp, default=cias_disp, key="f_cia")



def render_sidebar_exportacao() -> None:
    st.markdown('<div class="section-header">Exportar</div>', unsafe_allow_html=True)
    df_all = st.session_state.df_base

    if df_all.empty:
        return

    csv_bytes = df_all.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
    st.download_button(
        "Baixar CSV Completo",
        data=csv_bytes,
        file_name="base_atrasos_completa.csv",
        mime="text/csv",
        use_container_width=True,
    )


# -----------------------------------------------------------------------------
# Filtros e indicadores
# -----------------------------------------------------------------------------

def aplicar_filtros(df_all: pd.DataFrame) -> pd.DataFrame:
    df = df_all.copy()
    df["_ano"] = df["data"].astype(str).str[-4:]

    if st.session_state.get("f_ano"):
        df = df[df["_ano"].isin(st.session_state.f_ano)]
    if st.session_state.get("f_data"):
        df = df[df["data"].isin(st.session_state.f_data)]
    if st.session_state.get("f_icao"):
        df = df[df["icao"].isin(st.session_state.f_icao)]
    if st.session_state.get("f_tipo"):
        df = df[df["tipo_ocorrencia"].isin(st.session_state.f_tipo)]
    if st.session_state.get("f_mov"):
        df = df[df["movimento"].isin(st.session_state.f_mov)]
    if st.session_state.get("f_cia"):
        df = df[df["companhia"].isin(st.session_state.f_cia)]

    return df



def render_kpis(df: pd.DataFrame) -> None:
    total = len(df)
    n_aeroportos = df["icao"].nunique()
    n_datas = df["data"].nunique()
    min_medio = df["minutos_motivo_1"].mean()
    n_cancelados = len(df[df["tipo_ocorrencia"].astype(str).str.upper() == "CANCELADO"])

    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Total de Ocorrencias", f"{total:,}".replace(",", "."))
    col2.metric("Aeroportos com Dados", n_aeroportos)
    col3.metric("Datas na Base", n_datas)
    col4.metric("Atraso Medio (min)", f"{min_medio:.0f}" if not pd.isna(min_medio) else "-")
    col5.metric("Cancelamentos", n_cancelados)


# -----------------------------------------------------------------------------
# Abas
# -----------------------------------------------------------------------------

def render_tab_visao_geral(df: pd.DataFrame) -> None:
    col_l, col_r = st.columns(2)

    with col_l:
        st.markdown("#### Ocorrencias por Tipo")
        df_tipo = df["tipo_ocorrencia"].value_counts().reset_index()
        df_tipo.columns = ["Tipo", "Quantidade"]
        fig = px.bar(
            df_tipo,
            x="Tipo",
            y="Quantidade",
            color="Tipo",
            color_discrete_sequence=["#58a6ff", "#f85149", "#d29922", "#3fb950"],
        )
        fig.update_layout(**PLOT_THEME, showlegend=False, margin=dict(t=20, b=20))
        fig.update_traces(marker_line_width=0)
        st.plotly_chart(fig, use_container_width=True)

    with col_r:
        st.markdown("#### Distribuicao por Movimento")
        df_mov = df["movimento"].value_counts().reset_index()
        df_mov.columns = ["Movimento", "Quantidade"]
        fig2 = px.pie(
            df_mov,
            names="Movimento",
            values="Quantidade",
            color_discrete_sequence=["#58a6ff", "#3fb950", "#d29922", "#f85149"],
            hole=0.45,
        )
        fig2.update_layout(**PLOT_THEME, margin=dict(t=20, b=20))
        fig2.update_traces(textfont_color="#e6edf3")
        st.plotly_chart(fig2, use_container_width=True)

    if df["data"].nunique() > 1:
        st.markdown("#### Evolucao Diaria de Ocorrencias")
        df_dia = df.groupby("data").size().reset_index(name="Ocorrencias")
        fig3 = px.line(
            df_dia,
            x="data",
            y="Ocorrencias",
            markers=True,
            color_discrete_sequence=["#58a6ff"],
        )
        fig3.update_layout(**PLOT_THEME, margin=dict(t=20, b=20))
        fig3.update_traces(line_width=2.5, marker_size=8)
        st.plotly_chart(fig3, use_container_width=True)

    st.markdown("#### Distribuicao dos Atrasos (minutos)")
    df_min = df["minutos_motivo_1"].dropna()
    if not df_min.empty:
        fig4 = px.histogram(
            df_min,
            nbins=30,
            color_discrete_sequence=["#58a6ff"],
            labels={"value": "Minutos de Atraso", "count": "Frequencia"},
        )
        fig4.update_layout(**PLOT_THEME, margin=dict(t=20, b=20), showlegend=False)
        fig4.update_traces(marker_line_width=0)
        st.plotly_chart(fig4, use_container_width=True)



def render_tab_por_aeroporto(df: pd.DataFrame) -> None:
    st.markdown("#### Ocorrencias por Aeroporto")
    df_aero = df.groupby(["icao", "tipo_ocorrencia"]).size().reset_index(name="Quantidade")
    fig_aero = px.bar(
        df_aero,
        x="icao",
        y="Quantidade",
        color="tipo_ocorrencia",
        barmode="stack",
        color_discrete_map={
            "ATRASO": "#58a6ff",
            "CANCELADO": "#f85149",
            "RETORNO": "#d29922",
            "ALTERNADO": "#bc8cff",
        },
        labels={"icao": "ICAO", "tipo_ocorrencia": "Tipo"},
    )
    fig_aero.update_layout(**PLOT_THEME, margin=dict(t=20, b=20))
    st.plotly_chart(fig_aero, use_container_width=True)

    st.markdown("#### Atraso Medio por Aeroporto (min - Motivo 1)")
    df_med = (
        df.groupby("icao")["minutos_motivo_1"]
        .mean()
        .dropna()
        .round(1)
        .sort_values(ascending=False)
        .reset_index()
    )
    df_med.columns = ["ICAO", "Atraso Medio (min)"]
    fig_med = px.bar(
        df_med,
        x="ICAO",
        y="Atraso Medio (min)",
        color="Atraso Medio (min)",
        color_continuous_scale=["#1f3a5f", "#58a6ff", "#f85149"],
    )
    fig_med.update_layout(**PLOT_THEME, margin=dict(t=20, b=20), showlegend=False)
    st.plotly_chart(fig_med, use_container_width=True)



def render_tab_por_companhia(df: pd.DataFrame) -> None:
    col_l, col_r = st.columns(2)

    with col_l:
        st.markdown("#### Top Companhias por Ocorrencias")
        df_cia = df["companhia"].value_counts().head(15).reset_index()
        df_cia.columns = ["Companhia", "Ocorrencias"]
        fig_cia = px.bar(
            df_cia,
            x="Ocorrencias",
            y="Companhia",
            orientation="h",
            color="Ocorrencias",
            color_continuous_scale=["#1f3a5f", "#58a6ff"],
        )
        fig_cia.update_layout(
            **PLOT_THEME,
            yaxis=dict(autorange="reversed"),
            margin=dict(t=20, b=20),
            showlegend=False,
        )
        st.plotly_chart(fig_cia, use_container_width=True)

    with col_r:
        st.markdown("#### Atraso Medio por Companhia (min)")
        df_cia_med = (
            df.groupby("companhia")["minutos_motivo_1"]
            .mean()
            .dropna()
            .round(1)
            .sort_values(ascending=False)
            .head(15)
            .reset_index()
        )
        df_cia_med.columns = ["Companhia", "Atraso Medio (min)"]
        fig_cia_med = px.bar(
            df_cia_med,
            x="Atraso Medio (min)",
            y="Companhia",
            orientation="h",
            color="Atraso Medio (min)",
            color_continuous_scale=["#1a3028", "#3fb950", "#f85149"],
        )
        fig_cia_med.update_layout(
            **PLOT_THEME,
            yaxis=dict(autorange="reversed"),
            margin=dict(t=20, b=20),
            showlegend=False,
        )
        st.plotly_chart(fig_cia_med, use_container_width=True)



def render_tab_motivos(df: pd.DataFrame) -> None:
    st.markdown("#### Motivos de Atraso - Frequencia (Motivo 1)")
    todos_motivos = pd.concat(
        [
            df["motivo_1"].dropna(),
            df["motivo_2"].dropna(),
            df["motivo_3"].dropna(),
        ]
    )
    df_motivos = todos_motivos.value_counts().head(20).reset_index()
    df_motivos.columns = ["Codigo", "Frequencia"]
    fig_mot = px.bar(
        df_motivos,
        x="Frequencia",
        y="Codigo",
        orientation="h",
        color="Frequencia",
        color_continuous_scale=["#1f3a5f", "#58a6ff", "#f85149"],
    )
    fig_mot.update_layout(
        **PLOT_THEME,
        yaxis=dict(autorange="reversed"),
        margin=dict(t=20, b=20),
        showlegend=False,
    )
    st.plotly_chart(fig_mot, use_container_width=True)

    st.markdown("#### Atraso Medio por Codigo de Motivo (min)")
    df_mot_med = (
        df.groupby("motivo_1")["minutos_motivo_1"]
        .agg(["mean", "count"])
        .reset_index()
        .rename(columns={"mean": "Media (min)", "count": "Qtd"})
        .dropna()
        .query("Qtd >= 2")
        .sort_values("Media (min)", ascending=False)
        .head(20)
    )
    df_mot_med["Media (min)"] = df_mot_med["Media (min)"].round(1)

    fig_mot2 = px.scatter(
        df_mot_med,
        x="Qtd",
        y="Media (min)",
        text="motivo_1",
        size="Qtd",
        color="Media (min)",
        color_continuous_scale=["#3fb950", "#d29922", "#f85149"],
        labels={"Qtd": "Frequencia", "Media (min)": "Atraso Medio (min)"},
    )
    fig_mot2.update_traces(textposition="top center", textfont_size=10)
    fig_mot2.update_layout(**PLOT_THEME, margin=dict(t=30, b=20), showlegend=False)
    st.plotly_chart(fig_mot2, use_container_width=True)



def render_tab_dados_brutos(df: pd.DataFrame) -> None:
    st.markdown(f"**{len(df):,} registros** apos filtros aplicados".replace(",", "."))

    busca = st.text_input("Buscar em qualquer campo", placeholder="ex: AZU, BSB, RA - 93 ...")
    df_show = df.copy()

    if busca:
        mask = df_show.apply(lambda col: col.astype(str).str.contains(busca, case=False, na=False)).any(axis=1)
        df_show = df_show[mask]
        st.caption(f"{len(df_show)} registros encontrados para '{busca}'")

    st.dataframe(
        df_show.reset_index(drop=True),
        use_container_width=True,
        height=500,
        column_config={
            "data": st.column_config.TextColumn("Data"),
            "icao": st.column_config.TextColumn("ICAO"),
            "aeroporto": st.column_config.TextColumn("Aeroporto"),
            "item": st.column_config.NumberColumn("Item"),
            "tipo_ocorrencia": st.column_config.TextColumn("Tipo"),
            "movimento": st.column_config.TextColumn("Movimento"),
            "motivo_1": st.column_config.TextColumn("Motivo 1"),
            "minutos_motivo_1": st.column_config.NumberColumn("Min 1", format="%d min"),
            "motivo_2": st.column_config.TextColumn("Motivo 2"),
            "minutos_motivo_2": st.column_config.NumberColumn("Min 2", format="%d min"),
            "motivo_3": st.column_config.TextColumn("Motivo 3"),
            "minutos_motivo_3": st.column_config.NumberColumn("Min 3", format="%d min"),
            "companhia": st.column_config.TextColumn("Cia"),
            "numero_voo": st.column_config.TextColumn("Voo"),
            "equipamento": st.column_config.TextColumn("Equip."),
            "origem_destino": st.column_config.TextColumn("Orig/Dest"),
            "af_aeroporto": st.column_config.TextColumn("AF Aeroporto"),
        },
    )

    csv_filtrado = df_show.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
    st.download_button(
        "Baixar selecao como CSV",
        data=csv_filtrado,
        file_name="atrasos_filtrados.csv",
        mime="text/csv",
    )


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------

def render_footer() -> None:
    st.divider()
    st.markdown(
        "<div style='text-align:center; color:#8b949e; font-size:0.75rem; font-family:Space Mono, monospace;'>"
        "Base de Dados de Atrasos Aeroportuarios - ABR / ANAC - "
        + datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
        + "</div>",
        unsafe_allow_html=True,
    )



def main() -> None:
    configurar_pagina()
    inicializar_session_state()
    render_sidebar()

    st.markdown("# Base de Dados - Atrasos Aeroportuarios")
    df_all = st.session_state.df_base

    if df_all.empty:
        st.info("Nenhum dado carregado. Use o painel lateral para importar arquivos RD_*.xlsx.")
        st.stop()

    df = aplicar_filtros(df_all)
    render_kpis(df)
    st.divider()

    tab_vis, tab_aero, tab_cia, tab_motivos, tab_dados = st.tabs(
        ["Visao Geral", "Por Aeroporto", "Por Companhia", "Motivos", "Dados Brutos"]
    )

    with tab_vis:
        render_tab_visao_geral(df)
    with tab_aero:
        render_tab_por_aeroporto(df)
    with tab_cia:
        render_tab_por_companhia(df)
    with tab_motivos:
        render_tab_motivos(df)
    with tab_dados:
        render_tab_dados_brutos(df)

    render_footer()


if __name__ == "__main__":
    main()
